import os
import subprocess
import glob
import pandas as pd
import sys

# ================= USER CONFIGURATION =================
MULTIWFN_PATH = r"C:\Users\Admin\Downloads\Multiwfn_3.8_dev_bin_Win64\Multiwfn_3.8_dev_bin_Win64\Multiwfn.exe"
STATES_TO_CHECK = 3
# ======================================================

def get_homo_index(fchk_path):
    try:
        with open(fchk_path, 'r') as f:
            for line in f:
                if "Number of alpha electrons" in line:
                    return int(line.split()[-1])
    except:
        return None
    return None

def get_label(orb_idx, homo_idx):
    diff = orb_idx - homo_idx
    if diff == 0: return "HOMO"
    if diff == 1: return "LUMO"
    if diff < 0: return f"H{diff}" 
    if diff > 1: return f"L+{diff-1}"
    return str(orb_idx)

def parse_log_last_geometry(log_path, max_states):
    if not os.path.exists(log_path): return [], None

    with open(log_path, 'r') as f:
        content = f.read()

    # Split by TD-DFT header to get steps
    blocks = content.split("Excitation energies and oscillator strengths")
    if len(blocks) < 2: return [], None
    
    # Take ONLY the last block (Final Geometry)
    final_block = blocks[-1]
    lines = final_block.splitlines()

    states_data = []
    opt_state = None 
    
    current_state_info = None
    current_max_coeff = 0.0
    
    for line in lines:
        if "This state for optimization" in line and current_state_info:
            opt_state = current_state_info['state']

        if "Excited State" in line:
            if current_state_info: states_data.append(current_state_info)
            try:
                parts = line.split()
                state_idx = int(parts[2].replace(":", ""))
                f_val = float(line.split("f=")[1].strip().split()[0])
                current_state_info = {'state': state_idx, 'f': f_val, 'pair': None}
                current_max_coeff = 0.0
            except:
                current_state_info = None

        if current_state_info and "->" in line:
            try:
                t_parts = line.split()
                start = int(t_parts[0])
                end = int(t_parts[2])
                coeff = abs(float(t_parts[3]))
                if coeff > current_max_coeff:
                    current_max_coeff = coeff
                    current_state_info['pair'] = (start, end)
            except:
                pass

    if current_state_info: states_data.append(current_state_info)
    return states_data, opt_state

def run_calculation():
    fchk_files = sorted(glob.glob("*.fchk"))
    if not fchk_files:
        print("No .fchk files found!")
        return

    print(f"Found {len(fchk_files)} files.")
    print(f"Analyzing FINAL Geometry (Last 'Excitation energies' block).")
    print(f"Output: results_opt_merged.xlsx")
    print("="*100)
    print(f"{'Filename':<35} | {'St':<2} | {'Interest':<8} | {'f-val':<8} | {'Trans.':<7} | {'Character':<15} | {'dCT(A)':<8}")
    print("="*100)

    all_data_rows = []
    summary_rows = []

    # --- 1. DATA COLLECTION ---
    for file in fchk_files:
        base = os.path.splitext(file)[0]
        log_file = base + ".log"
        if not os.path.exists(log_file): log_file = base + ".out"
        
        homo_idx = get_homo_index(file)
        if not homo_idx: continue

        states_list, opt_state = parse_log_last_geometry(log_file, STATES_TO_CHECK)
        if not states_list: continue

        for i, data in enumerate(states_list):
            if data['state'] > STATES_TO_CHECK: continue

            s_idx = data['state']
            f_val = data['f']
            pair = data['pair']
            if not pair: continue

            start, end = pair
            lbl_start = get_label(start, homo_idx)
            lbl_end = get_label(end, homo_idx)
            char_label = f"{lbl_start}->{lbl_end}"
            orb_input = f"{start},{end}"

            # Multiwfn Calc
            cmds = ["100", "11", orb_input, "n", "0,0", "0", "q"]
            try:
                process = subprocess.run(
                    [MULTIWFN_PATH, file],
                    input="\n".join(cmds), text=True, capture_output=True,
                    encoding='utf-8', errors='ignore'
                )
                dist = "---"
                for line in process.stdout.splitlines():
                    if "Centroid distance" in line:
                        dist = float(line.split(":")[-1].strip().split()[0])
                        break
                
                is_interest = (s_idx == opt_state)
                interest_marker = "YES" if is_interest else ""
                
                # Print to console (Only print filename on first row of block)
                fname_print = file if i == 0 else "" 
                print(f"{fname_print:<35} | {s_idx:<2} | {interest_marker:<8} | {f_val:<8.4f} | {orb_input:<7} | {char_label:<15} | {dist:<8}")
                
                row_dict = {
                    "Filename": file,
                    "State": s_idx,
                    "Interest": interest_marker,
                    "f-value": f_val,
                    "Orbitals": orb_input,
                    "Character": char_label,
                    "dCT (Ang)": dist,
                    "is_bold": is_interest
                }
                all_data_rows.append(row_dict)
                if is_interest:
                    summary_rows.append(row_dict)

            except Exception as e:
                print(f"Error: {e}")
        print("-" * 100)

    # --- 2. EXCEL SAVING AND MERGING ---
    output_file = "results_opt_merged.xlsx"
    df_all = pd.DataFrame(all_data_rows)
    df_summary = pd.DataFrame(summary_rows)

    if "is_bold" in df_summary.columns:
        df_summary = df_summary.drop(columns=["is_bold"])

    try:
        import openpyxl 
        from openpyxl.styles import Alignment, Font

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Write sheets
            df_clean = df_all.drop(columns=["is_bold"])
            df_clean.to_excel(writer, sheet_name='All_States', index=False)
            df_summary.to_excel(writer, sheet_name='Summary_Opt_Root', index=False)
            
            # --- POST-PROCESSING ---
            workbook = writer.book
            ws = writer.sheets['All_States']
            
            # A. Bold Logic
            bold_font = Font(bold=True)
            for idx, row_data in df_all.iterrows():
                if row_data['is_bold']:
                    excel_row = idx + 2 # Header is 1
                    # Bold columns B to G (State to dCT)
                    # We skip column A (Filename) because merging messes with individual cell styles
                    for col in range(2, len(df_clean.columns) + 1):
                        cell = ws.cell(row=excel_row, column=col)
                        cell.font = bold_font

            # B. Merge Logic for Filename (Column A)
            # We iterate down column A. If A[i] == A[i-1], we extend the merge range.
            start_row = 2
            current_val = ws.cell(row=2, column=1).value
            max_row = ws.max_row

            # Iterate from row 3 to end+1
            for row in range(3, max_row + 2):
                val = ws.cell(row=row, column=1).value
                
                if val != current_val:
                    # The value changed, merge the previous block
                    if row - start_row > 1:
                        ws.merge_cells(start_row=start_row, start_column=1, end_row=row-1, end_column=1)
                        
                        # Style the merged cell (Top-Left is the anchor)
                        top_left = ws.cell(row=start_row, column=1)
                        top_left.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Reset for next block
                    start_row = row
                    current_val = val

            # Adjust column width for visibility
            ws.column_dimensions['A'].width = 50

        print(f"\nSuccess! Merged Excel saved as: {output_file}")

    except Exception as e:
        print(f"\nError creating Excel: {e}")
        df_all.drop(columns=["is_bold"]).to_csv("results_backup.csv", index=False)

if __name__ == "__main__":
    run_calculation()
